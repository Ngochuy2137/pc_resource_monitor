#!/usr/bin/env python3
"""
Log CPU and GPU usage to an Excel file on Jetson (or any Linux host with psutil).

CPU sampling follows main_controller/observer.py:
  psutil.cpu_percent(interval=0, percpu=True)
  psutil.virtual_memory().used

GPU load is read from Jetson sysfs (value / 10 = percent).
"""
from __future__ import annotations

import argparse
import os
import statistics
import subprocess
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

try:
    import psutil
except ImportError:
    print("Install dependencies: pip install -r requirements.txt", file=sys.stderr)
    raise SystemExit(1)

try:
    from openpyxl import Workbook
except ImportError:
    print("Install dependencies: pip install -r requirements.txt", file=sys.stderr)
    raise SystemExit(1)

DEFAULT_GPU_LOAD_PATHS = (
    "/sys/devices/platform/bus@0/17000000.gpu/load",
    "/sys/class/devfreq/17000000.gpu/device/load",
)
DEFAULT_PID_REFRESH_INTERVAL_S = 2.0
DEFAULT_SPIKE_TOP_N = 5
DEFAULT_SPIKE_SAMPLE_S = 0.2
DEFAULT_SPIKE_COOLDOWN_S = 5.0
PACKAGE_DIR = Path(__file__).resolve().parent
LOGS_DIR = PACKAGE_DIR / "logs"


@dataclass(frozen=True)
class ProcessCpuSnapshot:
    rank: int
    pid: int
    name: str
    cpu_pct: float
    ram_mb: float
    cmdline: str


@dataclass(frozen=True)
class CpuSpikeEvent:
    timestamp: str
    system_cpu_avg_pct: float
    system_cores_above: int
    top_processes: tuple[ProcessCpuSnapshot, ...]


def find_gpu_load_path() -> Path | None:
    for path_str in DEFAULT_GPU_LOAD_PATHS:
        path = Path(path_str)
        if path.is_file():
            return path

    devfreq = Path("/sys/class/devfreq")
    if devfreq.is_dir():
        for entry in devfreq.iterdir():
            name_file = entry / "device/of_node/name"
            if name_file.is_file():
                name = name_file.read_text(encoding="utf-8", errors="ignore").strip("\x00")
                if name == "gpu":
                    load_file = entry / "device/load"
                    if load_file.is_file():
                        return load_file
    return None


def read_gpu_percent(gpu_load_path: Path | None) -> float | None:
    if gpu_load_path is None:
        return None
    try:
        raw = gpu_load_path.read_text(encoding="utf-8").strip()
        return float(raw) / 10.0
    except (OSError, ValueError):
        return None


def get_cpu_count() -> int:
    return psutil.cpu_count() or 1


def normalize_process_cpu_to_system_scale(process_cpu_pct: float) -> float:
    """Convert psutil process CPU (% of one core) to system-avg scale."""
    return process_cpu_pct / get_cpu_count()


def get_cpu_per_core() -> list[float]:
    return list(psutil.cpu_percent(interval=0, percpu=True))


def get_system_ram_gb() -> float:
    return psutil.virtual_memory().used / (1024**3)


def count_cores_above(per_core: list[float], threshold: float) -> int:
    return sum(1 for value in per_core if value > threshold)


def find_pids_by_cmdline(fragment: str, *, include_self: bool = True) -> list[int]:
    fragment_lower = fragment.lower()
    own_pid = os.getpid()
    pids: list[int] = []
    for proc in psutil.process_iter(["pid", "cmdline", "exe"]):
        pid = proc.info["pid"]
        if pid == own_pid and not include_self:
            continue
        cmdline = " ".join(proc.info["cmdline"] or [])
        if fragment_lower not in cmdline.lower():
            continue
        exe = (proc.info["exe"] or "").lower()
        if exe.endswith("/bash") or exe.endswith("/sh"):
            continue
        pids.append(pid)
    return pids


def normalize_ros2_node_name(node_name: str) -> str:
    return node_name.strip().lstrip("/")


def ros2_node_to_process_match(node_name: str) -> str:
    """ROS2 passes the logical node name in process cmdline as __node:=NAME."""
    return f"__node:={normalize_ros2_node_name(node_name)}"


SELF_PROCESS_ALIAS = "self"


def is_self_process_alias(name: str) -> bool:
    return normalize_ros2_node_name(name).lower() == SELF_PROCESS_ALIAS


def match_fragments_for_target(name: str) -> list[str]:
    """Cmdline substrings: ROS2 __node:=NAME, or script NAME.py (not bare NAME)."""
    normalized = normalize_ros2_node_name(name)
    fragments = [ros2_node_to_process_match(normalized)]
    if not normalized.endswith(".py"):
        fragments.append(f"{normalized}.py")
    return fragments


def find_pids_by_ros2_node(node_name: str) -> list[int]:
    return find_pids_for_targets([node_name]).get(node_name, [])


def find_pids_for_targets(node_names: list[str]) -> dict[str, list[int]]:
    """Resolve target PIDs in one process scan (ROS2 nodes, scripts, or self)."""
    if not node_names:
        return {}

    ros2_fragments = {
        name: ros2_node_to_process_match(name).lower() for name in node_names
    }
    script_fragments = {
        name: f"{normalize_ros2_node_name(name)}.py".lower()
        for name in node_names
        if not normalize_ros2_node_name(name).endswith(".py")
    }
    pids_by_name = {name: [] for name in node_names}
    ros2_pids_by_name = {name: [] for name in node_names}
    script_pids_by_name = {name: [] for name in node_names}
    own_pid = os.getpid()

    for name in node_names:
        if is_self_process_alias(name):
            pids_by_name[name] = [own_pid]

    for proc in psutil.process_iter(["pid", "cmdline", "exe"]):
        pid = proc.info["pid"]
        cmdline = " ".join(proc.info["cmdline"] or [])
        cmdline_lower = cmdline.lower()
        exe = (proc.info["exe"] or "").lower()
        if exe.endswith("/bash") or exe.endswith("/sh"):
            continue
        for name in node_names:
            if is_self_process_alias(name):
                continue
            if ros2_fragments[name] in cmdline_lower:
                if pid not in ros2_pids_by_name[name]:
                    ros2_pids_by_name[name].append(pid)
            script_fragment = script_fragments.get(name)
            if script_fragment and script_fragment in cmdline_lower:
                if pid not in script_pids_by_name[name]:
                    script_pids_by_name[name].append(pid)

    for name in node_names:
        if is_self_process_alias(name):
            continue
        if ros2_pids_by_name[name]:
            pids_by_name[name] = ros2_pids_by_name[name]
        else:
            pids_by_name[name] = script_pids_by_name[name]

    return pids_by_name


def find_pids_for_ros2_nodes(node_names: list[str]) -> dict[str, list[int]]:
    return find_pids_for_targets(node_names)


class TargetPidTracker:
    """Cache ROS2 node PIDs; refresh periodically or after a stale PID."""

    def __init__(
        self,
        target_processes: list[str],
        refresh_interval_s: float = DEFAULT_PID_REFRESH_INTERVAL_S,
    ) -> None:
        self.target_processes = target_processes
        self.refresh_interval_s = refresh_interval_s
        self.pids_by_name: dict[str, list[int]] = {
            name: [] for name in target_processes
        }
        self.last_refresh = 0.0
        self._force_refresh = True

    def refresh(self) -> dict[str, list[int]]:
        self.pids_by_name = find_pids_for_ros2_nodes(self.target_processes)
        self.last_refresh = time.monotonic()
        self._force_refresh = False
        return self.pids_by_name

    def get_pids(self, now: float) -> dict[str, list[int]]:
        if self._force_refresh or (now - self.last_refresh) >= self.refresh_interval_s:
            self.refresh()
        return self.pids_by_name

    def invalidate(self) -> None:
        self._force_refresh = True


def normalize_target_processes(names: list[str]) -> list[str]:
    return [normalize_ros2_node_name(name) for name in names if name.strip()]


def build_columns(threshold: float, target_processes: list[str]) -> list[str]:
    cores_col = f"system_cores_above_{threshold:g}pct"
    columns = [
        "timestamp",
        "system_cpu %",
        "system_cpu_AVERAGE",
        "system_gpu %",
        "system_gpu_AVERAGE",
        "system_ram_gb",
        cores_col,
    ]
    for name in target_processes:
        columns.append(f"{name}_cpu %")
        columns.append(f"{name}_ram_gb")
    return columns


# Per-PID baseline: (cpu_time_seconds, monotonic_timestamp)
ProcessCpuBaseline = tuple[float, float]


def read_process_cpu_pct(
    pid: int,
    proc: psutil.Process,
    baselines: dict[int, ProcessCpuBaseline],
    min_read_interval_s: float,
) -> float | None:
    """CPU usage (% of one core) from cpu_times delta; None = skip this sample."""
    read_now = time.monotonic()
    times = proc.cpu_times()
    cpu_sec = times.user + times.system

    if pid not in baselines:
        baselines[pid] = (cpu_sec, read_now)
        return None

    prev_cpu, prev_t = baselines[pid]
    delta_t = read_now - prev_t
    if delta_t < min_read_interval_s:
        return None

    delta_cpu = cpu_sec - prev_cpu
    if delta_cpu < 0:
        baselines[pid] = (cpu_sec, read_now)
        return None

    baselines[pid] = (cpu_sec, read_now)
    raw_pct = (delta_cpu / delta_t) * 100.0
    # One process cannot exceed 100% x cpu_count on the per-core scale.
    if raw_pct > get_cpu_count() * 100.0:
        return None
    return raw_pct


def sample_single_pid(
    pid: int,
    process_cache: dict[int, psutil.Process],
    cpu_baselines: dict[int, ProcessCpuBaseline],
    min_read_interval_s: float,
) -> tuple[float | None, float, bool]:
    """Return (cpu_pct per-core scale or None, ram_gb, stale_pid)."""
    try:
        proc = process_cache.get(pid)
        if proc is None:
            proc = psutil.Process(pid)
            process_cache[pid] = proc

        cpu_pct = read_process_cpu_pct(pid, proc, cpu_baselines, min_read_interval_s)
        ram_gb = proc.memory_info().rss / (1024**3)
        return cpu_pct, ram_gb, False
    except (psutil.NoSuchProcess, psutil.AccessDenied):
        process_cache.pop(pid, None)
        cpu_baselines.pop(pid, None)
        return None, 0.0, True


def sample_process_resource(
    pids: list[int],
    process_cache: dict[int, psutil.Process],
    cpu_baselines: dict[int, ProcessCpuBaseline],
    min_read_interval_s: float,
    pid_samples: dict[int, tuple[float | None, float]] | None = None,
) -> tuple[float, float, bool]:
    cpu_total = 0.0
    ram_gb = 0.0
    stale_pid = False
    got_reading = False

    for pid in pids:
        if pid_samples is not None and pid in pid_samples:
            cpu_pct, pid_ram = pid_samples[pid]
            stale_pid = False
        else:
            cpu_pct, pid_ram, stale_pid = sample_single_pid(
                pid, process_cache, cpu_baselines, min_read_interval_s
            )
            if pid_samples is not None:
                pid_samples[pid] = (cpu_pct, pid_ram)

        if stale_pid:
            continue
        if cpu_pct is not None:
            cpu_total += cpu_pct
            got_reading = True
        ram_gb += pid_ram

    if not got_reading:
        cpu_total = 0.0
    return normalize_process_cpu_to_system_scale(cpu_total), ram_gb, stale_pid


def sample_targets_for_tick(
    target_processes: list[str],
    pids_by_name: dict[str, list[int]],
    process_cache: dict[int, psutil.Process],
    cpu_baselines: dict[int, ProcessCpuBaseline],
    min_read_interval_s: float,
) -> tuple[dict[str, tuple[float, float]], bool]:
    """Sample each unique PID once, then aggregate per target name."""
    unique_pids: list[int] = []
    seen: set[int] = set()
    for name in target_processes:
        for pid in pids_by_name.get(name, []):
            if pid not in seen:
                seen.add(pid)
                unique_pids.append(pid)

    pid_samples: dict[int, tuple[float | None, float]] = {}
    stale_pid = False
    for pid in unique_pids:
        cpu_pct, ram_gb, pid_stale = sample_single_pid(
            pid, process_cache, cpu_baselines, min_read_interval_s
        )
        pid_samples[pid] = (cpu_pct, ram_gb)
        stale_pid = stale_pid or pid_stale

    results: dict[str, tuple[float, float]] = {}
    for name in target_processes:
        pids = pids_by_name.get(name, [])
        cpu_pct, ram_gb, name_stale = sample_process_resource(
            pids,
            process_cache,
            cpu_baselines,
            min_read_interval_s,
            pid_samples=pid_samples,
        )
        results[name] = (cpu_pct, ram_gb)
        stale_pid = stale_pid or name_stale
    return results, stale_pid


def init_cpu_baselines(
    pids: list[int],
    process_cache: dict[int, psutil.Process],
    cpu_baselines: dict[int, ProcessCpuBaseline],
    now: float,
) -> None:
    for pid in pids:
        try:
            proc = process_cache.get(pid)
            if proc is None:
                proc = psutil.Process(pid)
                process_cache[pid] = proc
            times = proc.cpu_times()
            cpu_baselines[pid] = (times.user + times.system, now)
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            process_cache.pop(pid, None)
            cpu_baselines.pop(pid, None)


def evict_untracked_pids(
    pids_by_name: dict[str, list[int]],
    process_cache: dict[int, psutil.Process],
    cpu_baselines: dict[int, ProcessCpuBaseline],
) -> None:
    tracked = set()
    for pids in pids_by_name.values():
        tracked.update(pids)
    for pid in list(process_cache):
        if pid not in tracked:
            process_cache.pop(pid, None)
            cpu_baselines.pop(pid, None)


ProcessCpuTimesSample = tuple[float, str, str, float]


def collect_process_cpu_times_samples(
    *,
    exclude_self: bool = True,
) -> dict[int, ProcessCpuTimesSample]:
    """One-shot sample of per-process cpu_times for spike ranking."""
    own_pid = os.getpid()
    samples: dict[int, ProcessCpuTimesSample] = {}
    for proc in psutil.process_iter(["pid", "name", "cmdline"]):
        pid = proc.info["pid"]
        if exclude_self and pid == own_pid:
            continue
        try:
            times = proc.cpu_times()
            name = proc.info["name"] or "?"
            cmdline_parts = proc.info["cmdline"] or []
            cmdline = " ".join(cmdline_parts)[:200]
            ram_mb = proc.memory_info().rss / (1024**2)
            samples[pid] = (times.user + times.system, name, cmdline, ram_mb)
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            continue
    return samples


def snapshot_top_processes_by_cpu_delta(
    top_n: int,
    sample_interval_s: float,
    *,
    exclude_self: bool = True,
) -> list[ProcessCpuSnapshot]:
    """
    Rank processes by CPU over a short window. Only called on CPU spikes,
    not on every monitor tick.
    """
    if top_n <= 0 or sample_interval_s <= 0:
        return []

    first = collect_process_cpu_times_samples(exclude_self=exclude_self)
    time.sleep(sample_interval_s)
    second = collect_process_cpu_times_samples(exclude_self=exclude_self)

    cpu_count = get_cpu_count()
    ranked: list[tuple[float, int, str, float, str]] = []
    for pid, (cpu_sec, name, cmdline, ram_mb) in second.items():
        first_sample = first.get(pid)
        if first_sample is None:
            continue
        delta_cpu = cpu_sec - first_sample[0]
        if delta_cpu < 0:
            continue
        cpu_pct = (delta_cpu / sample_interval_s) * 100.0 / cpu_count
        if cpu_pct <= 0.01:
            continue
        ranked.append((cpu_pct, pid, name, ram_mb, cmdline))

    ranked.sort(key=lambda item: item[0], reverse=True)
    return [
        ProcessCpuSnapshot(
            rank=index + 1,
            pid=pid,
            name=name,
            cpu_pct=cpu_pct,
            ram_mb=ram_mb,
            cmdline=cmdline,
        )
        for index, (cpu_pct, pid, name, ram_mb, cmdline) in enumerate(ranked[:top_n])
    ]


def snapshot_top_processes_via_pidstat(
    top_n: int,
    sample_interval_s: float,
) -> list[ProcessCpuSnapshot] | None:
    """Optional fast path using pidstat (sysstat). Returns None if unavailable."""
    if top_n <= 0:
        return []

    count = max(1, int(round(sample_interval_s)))
    try:
        completed = subprocess.run(
            ["pidstat", "-u", str(count), "1"],
            capture_output=True,
            text=True,
            check=False,
            timeout=sample_interval_s + 5.0,
        )
    except (FileNotFoundError, subprocess.TimeoutExpired, OSError):
        return None

    if completed.returncode != 0 or not completed.stdout.strip():
        return None

    own_pid = os.getpid()
    totals: dict[int, tuple[float, str]] = {}
    for line in completed.stdout.splitlines():
        if not line.startswith("  "):
            continue
        parts = line.split()
        if len(parts) < 9 or parts[0] != "Average:":
            continue
        try:
            pid = int(parts[2])
            if pid == own_pid:
                continue
            cpu_pct = float(parts[7])
            command = " ".join(parts[9:])[:200]
        except (ValueError, IndexError):
            continue
        prev_cpu = totals.get(pid, (command, 0.0))[1]
        totals[pid] = (command, prev_cpu + cpu_pct)

    if not totals:
        return None

    ranked = sorted(totals.items(), key=lambda item: item[1][1], reverse=True)
    snapshots: list[ProcessCpuSnapshot] = []
    for index, (pid, (cmdline, cpu_pct)) in enumerate(ranked[:top_n]):
        name = cmdline.split()[0] if cmdline else "?"
        ram_mb = 0.0
        try:
            ram_mb = psutil.Process(pid).memory_info().rss / (1024**2)
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            pass
        snapshots.append(
            ProcessCpuSnapshot(
                rank=index + 1,
                pid=pid,
                name=name,
                cpu_pct=cpu_pct,
                ram_mb=ram_mb,
                cmdline=cmdline,
            )
        )
    return snapshots


def capture_cpu_spike_top_processes(
    top_n: int,
    sample_interval_s: float,
) -> list[ProcessCpuSnapshot]:
    pidstat_result = snapshot_top_processes_via_pidstat(top_n, sample_interval_s)
    if pidstat_result is not None:
        return pidstat_result
    return snapshot_top_processes_by_cpu_delta(top_n, sample_interval_s)


class CpuSpikeDetector:
    """Trigger a one-shot top-process snapshot only when CPU crosses a threshold."""

    def __init__(
        self,
        threshold_pct: float,
        top_n: int,
        sample_interval_s: float,
        cooldown_s: float,
        *,
        enabled: bool = True,
    ) -> None:
        self.threshold_pct = threshold_pct
        self.top_n = top_n
        self.sample_interval_s = sample_interval_s
        self.cooldown_s = cooldown_s
        self.enabled = enabled and threshold_pct > 0
        self._above_threshold = False
        self._last_capture = 0.0

    def maybe_capture(
        self,
        *,
        timestamp: str,
        system_cpu_avg_pct: float,
        system_cores_above: int,
        now: float,
    ) -> CpuSpikeEvent | None:
        if not self.enabled:
            return None

        is_spike = system_cpu_avg_pct >= self.threshold_pct
        if not is_spike:
            self._above_threshold = False
            return None

        rising_edge = not self._above_threshold
        cooldown_elapsed = (now - self._last_capture) >= self.cooldown_s
        if not rising_edge and not cooldown_elapsed:
            self._above_threshold = True
            return None

        self._above_threshold = True
        self._last_capture = now
        top_processes = tuple(
            capture_cpu_spike_top_processes(self.top_n, self.sample_interval_s)
        )
        return CpuSpikeEvent(
            timestamp=timestamp,
            system_cpu_avg_pct=system_cpu_avg_pct,
            system_cores_above=system_cores_above,
            top_processes=top_processes,
        )


def format_spike_event(event: CpuSpikeEvent) -> str:
    lines = [
        (
            f"CPU spike: system_cpu={event.system_cpu_avg_pct:.1f}% "
            f"cores_above={event.system_cores_above} @ {event.timestamp}"
        )
    ]
    if not event.top_processes:
        lines.append("  (no process snapshot available)")
        return "\n".join(lines)

    for proc in event.top_processes:
        lines.append(
            f"  #{proc.rank} pid={proc.pid} cpu={proc.cpu_pct:.1f}% "
            f"ram={proc.ram_mb:.0f}MB {proc.name}: {proc.cmdline}"
        )
    return "\n".join(lines)


SPIKE_SHEET_COLUMNS = [
    "spike_timestamp",
    "system_cpu %",
    "system_cores_above",
    "rank",
    "pid",
    "name",
    "cpu %",
    "ram_mb",
    "cmdline",
]


def append_spike_event_rows(
    sheet,
    event: CpuSpikeEvent,
    *,
    separator_before: bool = False,
) -> None:
    if separator_before:
        sheet.append([None] * len(SPIKE_SHEET_COLUMNS))

    if not event.top_processes:
        sheet.append(
            [
                event.timestamp,
                round(event.system_cpu_avg_pct, 2),
                event.system_cores_above,
                None,
                None,
                None,
                None,
                None,
                None,
            ]
        )
        return

    for proc in event.top_processes:
        sheet.append(
            [
                event.timestamp,
                round(event.system_cpu_avg_pct, 2),
                event.system_cores_above,
                proc.rank,
                proc.pid,
                proc.name,
                round(proc.cpu_pct, 2),
                round(proc.ram_mb, 1),
                proc.cmdline,
            ]
        )


def format_elapsed_prefix(elapsed_s: float, duration_s: float) -> str:
    if duration_s > 0:
        return f"[{elapsed_s:.1f}/{duration_s:g}]"
    return f"[{elapsed_s:.1f}s]"


def format_status_line(
    row: dict[str, object],
    elapsed_s: float,
    duration_s: float,
    system_cores_col: str,
    target_processes: list[str],
) -> str:
    gpu_pct = row["system_gpu %"]
    gpu_text = "n/a" if gpu_pct is None else f"{float(gpu_pct):.1f}"
    line = (
        f"{format_elapsed_prefix(elapsed_s, duration_s)} "
        f"system_cpu={float(row['system_cpu %']):.1f}% "
        f"system_gpu={gpu_text}% "
        f"system_ram={float(row['system_ram_gb']):.3f}GB "
        f"{system_cores_col}={row[system_cores_col]}"
    )
    for name in target_processes:
        line += (
            f" {name}_cpu={float(row[f'{name}_cpu %']):.1f}%"
            f" {name}_ram={float(row[f'{name}_ram_gb']):.3f}GB"
        )
    return line


def print_status_line(line: str) -> None:
    print("\r" + line, end="", flush=True)


def default_output_path() -> Path:
    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return LOGS_DIR / f"cpu_gpu_monitor_{stamp}.xlsx"


def resolve_output_path(user_path: Path | None) -> Path:
    if user_path is None:
        return default_output_path()
    if user_path.is_absolute():
        user_path.parent.mkdir(parents=True, exist_ok=True)
        return user_path
    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    return LOGS_DIR / user_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Monitor CPU/GPU usage and save samples to an Excel file."
    )
    parser.add_argument(
        "--hz",
        type=float,
        default=10.0,
        help="Sampling rate in Hz (default: 10)",
    )
    parser.add_argument(
        "-x",
        "--threshold",
        type=float,
        default=80.0,
        help="CPU core usage threshold in percent for the core-count column (default: 80)",
    )
    parser.add_argument(
        "--duration",
        type=float,
        default=0.0,
        help="Run duration in seconds (0 = until Ctrl+C, default: 0)",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Output .xlsx path (default: logs/cpu_gpu_monitor_<timestamp>.xlsx)",
    )
    parser.add_argument(
        "--target-processes",
        dest="target_processes",
        nargs="+",
        default=None,
        metavar="NAME",
        help=(
            "Process or ROS2 node name(s) to track. Matches __node:=NAME, "
            "cmdline substring, or NAME.py; use 'self' for this monitor process. "
            "Example: --target-processes micro_ros_agent monitor self"
        ),
    )
    parser.add_argument(
        "--spike-threshold",
        type=float,
        default=None,
        help=(
            "Enable spike capture when system_cpu >= this value (%%). "
            "Omit this flag to disable spike capture entirely."
        ),
    )
    parser.add_argument(
        "--spike-top-n",
        type=int,
        default=DEFAULT_SPIKE_TOP_N,
        help=(
            "Number of top CPU processes to record per spike "
            f"(default: {DEFAULT_SPIKE_TOP_N})"
        ),
    )
    parser.add_argument(
        "--spike-sample-s",
        type=float,
        default=DEFAULT_SPIKE_SAMPLE_S,
        help=(
            "CPU ranking window in seconds for spike snapshots "
            f"(default: {DEFAULT_SPIKE_SAMPLE_S})"
        ),
    )
    parser.add_argument(
        "--spike-cooldown-s",
        type=float,
        default=DEFAULT_SPIKE_COOLDOWN_S,
        help=(
            "Minimum seconds between spike snapshots during sustained high CPU "
            f"(default: {DEFAULT_SPIKE_COOLDOWN_S})"
        ),
    )
    return parser.parse_args()


def round_cell(column: str, value: object) -> object:
    if column == "timestamp":
        return value
    if value is None:
        return None
    if column.endswith(" %") or column.endswith("_AVERAGE"):
        return round(float(value), 2)
    if column.endswith("_gb"):
        return round(float(value), 3)
    if column.startswith("system_cores_above_"):
        return int(value)
    return value


def compute_system_averages(rows: list[dict[str, object]]) -> tuple[float, float | None]:
    cpu_avg = statistics.fmean(float(row["system_cpu %"]) for row in rows)
    gpu_values = [
        float(row["system_gpu %"])
        for row in rows
        if row["system_gpu %"] is not None
    ]
    gpu_avg = statistics.fmean(gpu_values) if gpu_values else None
    return cpu_avg, gpu_avg


def build_summary_row(columns: list[str], rows: list[dict[str, object]]) -> list[object]:
    cpu_avg, gpu_avg = compute_system_averages(rows)
    summary: list[object] = []
    for column in columns:
        if column == "timestamp":
            summary.append("SUMMARY")
            continue
        if column == "system_cpu_AVERAGE":
            summary.append(round(cpu_avg, 2))
            continue
        if column == "system_gpu_AVERAGE":
            summary.append(round(gpu_avg, 2) if gpu_avg is not None else None)
            continue

        values = [row[column] for row in rows]
        if column == "system_gpu %":
            gpu_values = [float(value) for value in values if value is not None]
            summary.append(
                round(statistics.fmean(gpu_values), 2) if gpu_values else None
            )
        elif column.startswith("system_cores_above_"):
            summary.append(sum(int(value) for value in values))
        elif column.endswith(" %") or column.endswith("_gb"):
            summary.append(
                round(statistics.fmean(float(value) for value in values), 3)
                if column.endswith("_gb")
                else round(statistics.fmean(float(value) for value in values), 2)
            )
        else:
            summary.append(None)
    return summary


def write_run_params_sheet(
    workbook,
    args: argparse.Namespace,
    target_processes: list[str],
    output_path: Path,
    spike_enabled: bool,
) -> None:
    duration_text = "until Ctrl+C" if args.duration <= 0 else f"{args.duration:g}s"
    params_sheet = workbook.create_sheet("run_params", 0)
    params_sheet.append(["parameter", "value"])
    params_sheet.append(["--hz", args.hz])
    params_sheet.append(["--threshold", args.threshold])
    params_sheet.append(["--duration", duration_text])
    params_sheet.append(["--output", str(output_path)])
    params_sheet.append(
        [
            "--target-processes",
            ",".join(target_processes) if target_processes else "",
        ]
    )
    params_sheet.append(
        ["--spike-threshold", args.spike_threshold if spike_enabled else ""]
    )
    if spike_enabled:
        params_sheet.append(["--spike-top-n", args.spike_top_n])
        params_sheet.append(["--spike-sample-s", args.spike_sample_s])
        params_sheet.append(["--spike-cooldown-s", args.spike_cooldown_s])


def main() -> int:
    args = parse_args()
    if args.hz <= 0:
        print("--hz must be > 0", file=sys.stderr)
        return 1
    if not 0 <= args.threshold <= 100:
        print("--threshold must be between 0 and 100", file=sys.stderr)
        return 1

    spike_enabled = args.spike_threshold is not None
    if spike_enabled:
        if not 0 < args.spike_threshold <= 100:
            print("--spike-threshold must be between 0 and 100 (exclusive of 0)", file=sys.stderr)
            return 1
        if args.spike_top_n < 0:
            print("--spike-top-n must be >= 0", file=sys.stderr)
            return 1
        if args.spike_sample_s <= 0:
            print("--spike-sample-s must be > 0", file=sys.stderr)
            return 1
        if args.spike_cooldown_s < 0:
            print("--spike-cooldown-s must be >= 0", file=sys.stderr)
            return 1

    target_processes = (
        normalize_target_processes(args.target_processes)
        if args.target_processes
        else []
    )
    columns = build_columns(args.threshold, target_processes)
    system_cores_col = f"system_cores_above_{args.threshold:g}pct"

    interval = 1.0 / args.hz
    output_path = resolve_output_path(args.output)
    gpu_load_path = find_gpu_load_path()
    process_cache: dict[int, psutil.Process] = {}
    cpu_baselines: dict[int, ProcessCpuBaseline] = {}
    pid_tracker = TargetPidTracker(target_processes)
    min_read_interval_s = interval * 0.75
    spike_detector = CpuSpikeDetector(
        threshold_pct=args.spike_threshold or 0.0,
        top_n=args.spike_top_n,
        sample_interval_s=args.spike_sample_s,
        cooldown_s=args.spike_cooldown_s,
        enabled=spike_enabled,
    )
    spike_events: list[CpuSpikeEvent] = []

    if gpu_load_path is None:
        print("Warning: GPU load sysfs not found; GPU column will be empty.", file=sys.stderr)
    else:
        print(f"GPU load path: {gpu_load_path}")

    psutil.cpu_percent(interval=0, percpu=True)
    prime_time = time.monotonic()
    initial_pids = pid_tracker.refresh()
    for name in target_processes:
        init_cpu_baselines(
            initial_pids.get(name, []), process_cache, cpu_baselines, prime_time
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "monitor"
    write_run_params_sheet(
        workbook, args, target_processes, output_path, spike_enabled
    )
    sheet.append(columns)
    spike_sheet = None
    if spike_detector.enabled:
        spike_sheet = workbook.create_sheet("cpu_spikes")
        spike_sheet.append(SPIKE_SHEET_COLUMNS)

    rows: list[dict[str, object]] = []
    start = time.monotonic()
    next_sample = start + interval
    sample_count = 0

    print(
        f"Sampling at {args.hz:g} Hz, threshold={args.threshold:g}%%, "
        f"cpu_count={get_cpu_count()}, output={output_path}"
    )
    if target_processes:
        print(
            "Target process CPU uses the same scale as system_cpu_avg_pct "
            "(cpu_times delta / elapsed / cpu_count)."
        )
    for name in target_processes:
        pids = initial_pids.get(name, [])
        if pids:
            if is_self_process_alias(name):
                match_desc = f"alias '{SELF_PROCESS_ALIAS}' (this process)"
            else:
                match_desc = (
                    "cmdline fragments: "
                    + ", ".join(match_fragments_for_target(name))
                )
            print(f"Tracking '/{name}': {match_desc}, PIDs={pids}")
            if len(pids) > 1:
                print(
                    f"Warning: multiple PIDs matched '/{name}'; "
                    "CPU is summed across all matches.",
                    file=sys.stderr,
                )
        else:
            print(
                f"Warning: no process matched '/{name}'; "
                "its columns will stay empty until a match appears.",
                file=sys.stderr,
            )
    if target_processes:
        print(
            f"PID cache: refresh every {DEFAULT_PID_REFRESH_INTERVAL_S:g}s "
            "or immediately after node restart."
        )
        print(
            "Process CPU scale: 0-100% = share of total system CPU "
            f"(cpu_times delta / elapsed / {get_cpu_count()} cores)."
        )
    if spike_detector.enabled:
        print(
            f"CPU spike capture: threshold={args.spike_threshold:g}%%, "
            f"top_n={args.spike_top_n}, sample={args.spike_sample_s:g}s, "
            f"cooldown={args.spike_cooldown_s:g}s "
            "(full process scan only on spike, not every tick)."
        )

    try:
        while True:
            now = time.monotonic()
            if args.duration > 0 and (now - start) >= args.duration:
                break

            if now < next_sample:
                time.sleep(min(0.001, next_sample - now))
                continue

            per_core = get_cpu_per_core()
            elapsed_s = now - start
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
            row: dict[str, object] = {
                "timestamp": timestamp,
                "system_cpu %": statistics.fmean(per_core) if per_core else 0.0,
                "system_gpu %": read_gpu_percent(gpu_load_path),
                "system_ram_gb": get_system_ram_gb(),
                system_cores_col: count_cores_above(per_core, args.threshold),
            }

            current_pids = pid_tracker.get_pids(now)
            evict_untracked_pids(current_pids, process_cache, cpu_baselines)
            target_samples, stale_pid = sample_targets_for_tick(
                target_processes,
                current_pids,
                process_cache,
                cpu_baselines,
                min_read_interval_s,
            )
            if stale_pid:
                pid_tracker.invalidate()
            for name in target_processes:
                cpu_pct, ram_gb = target_samples.get(name, (0.0, 0.0))
                row[f"{name}_cpu %"] = cpu_pct
                row[f"{name}_ram_gb"] = ram_gb

            spike_event = spike_detector.maybe_capture(
                timestamp=timestamp,
                system_cpu_avg_pct=float(row["system_cpu %"]),
                system_cores_above=int(row[system_cores_col]),
                now=now,
            )
            if spike_event is not None:
                spike_events.append(spike_event)
                print(f"\n{format_spike_event(spike_event)}", flush=True)
                if spike_sheet is not None:
                    append_spike_event_rows(
                        spike_sheet,
                        spike_event,
                        separator_before=len(spike_events) > 1,
                    )

            rows.append(row)
            sample_count += 1
            # Resync when behind; do not catch up missed ticks (burst breaks
            # process cpu_percent(interval=None) which needs real elapsed time).
            next_sample = now + interval

            if sample_count % int(max(args.hz, 1)) == 0:
                print_status_line(
                    format_status_line(
                        row, elapsed_s, args.duration, system_cores_col, target_processes
                    )
                )
    except KeyboardInterrupt:
        print("\nStopped by user.")
    else:
        print()

    if rows:
        cpu_avg, gpu_avg = compute_system_averages(rows)
        for row in rows:
            row["system_cpu_AVERAGE"] = cpu_avg
            row["system_gpu_AVERAGE"] = gpu_avg
            sheet.append([round_cell(column, row[column]) for column in columns])
        sheet.append(build_summary_row(columns, rows))

    workbook.save(output_path)
    print(f"Saved {len(rows)} samples to {output_path}")
    if spike_events:
        print(f"Recorded {len(spike_events)} CPU spike snapshot(s).")

    if rows:
        summary_row = build_summary_row(columns, rows)
        summary_map = dict(zip(columns, summary_row))
        print(
            "Summary row: "
            f"system_cpu_avg_pct={summary_map['system_cpu %']}% "
            f"system_ram_gb={summary_map['system_ram_gb']} "
            f"{system_cores_col}={summary_map[system_cores_col]}"
        )
        for name in target_processes:
            print(
                f"  {name}: cpu_avg_pct={summary_map[f'{name}_cpu %']}% "
                f"ram_gb={summary_map[f'{name}_ram_gb']}"
            )
        gpu_values = [
            float(row["system_gpu %"])
            for row in rows
            if row["system_gpu %"] is not None
        ]
        if gpu_values:
            print(
                f"System GPU stats: min={min(gpu_values):.1f}% "
                f"max={max(gpu_values):.1f}% "
                f"avg={statistics.fmean(gpu_values):.1f}%"
            )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
